import pandas as pd
import os
from openpyxl import Workbook
from openpyxl.utils.dataframe import dataframe_to_rows
from openpyxl.styles import Font, PatternFill, Alignment
from openpyxl.utils import get_column_letter

def ajustar_colunas_csv():
    uploads_dir = "uploads"
    
    # Lista dos arquivos CSV
    csv_files = [
        "filiais-regional.csv",
        "produtos-estoque.csv", 
        "vendas-varejo.csv",
        "vendas-atacado.csv"
    ]
    
    for csv_file in csv_files:
        file_path = os.path.join(uploads_dir, csv_file)
        
        if os.path.exists(file_path):
            print(f"Processando {csv_file}...")
            
            # Ler CSV
            df = pd.read_csv(file_path, sep=';')
            
            # Criar arquivo Excel com formatação
            wb = Workbook()
            ws = wb.active
            ws.title = csv_file.replace('.csv', '').replace('-', '_').title()
            
            # Adicionar dados
            for r in dataframe_to_rows(df, index=False, header=True):
                ws.append(r)
            
            # Estilizar cabeçalho
            header_fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
            header_font = Font(color="FFFFFF", bold=True)
            
            for cell in ws[1]:
                cell.fill = header_fill
                cell.font = header_font
                cell.alignment = Alignment(horizontal="center", vertical="center")
            
            # Ajustar largura das colunas automaticamente
            for column in ws.columns:
                max_length = 0
                column_letter = get_column_letter(column[0].column)
                
                for cell in column:
                    try:
                        if len(str(cell.value)) > max_length:
                            max_length = len(str(cell.value))
                    except:
                        pass
                
                # Adicionar margem extra e definir largura mínima
                adjusted_width = min(max(max_length + 2, 12), 50)
                ws.column_dimensions[column_letter].width = adjusted_width
            
            # Salvar como Excel
            excel_path = file_path.replace('.csv', '.xlsx')
            wb.save(excel_path)
            print(f"OK Criado {excel_path} com colunas ajustadas")
            
            # Também atualizar o CSV original com melhor formatação
            df.to_csv(file_path, sep=';', index=False, encoding='utf-8')
            
    print("\nTodos os arquivos foram processados!")
    print("Arquivos Excel criados com colunas totalmente visiveis")
    print("Use os arquivos .xlsx para melhor visualizacao")

if __name__ == "__main__":
    ajustar_colunas_csv()